import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Parameters
dt = 0.033
H = np.array([[1, 0]])
R = 0.1**2  # Measurement uncertainty (0.1 m)
SDg = 0.01 * 9.81  # Standard deviation of acceleration (1% of 9.81)

# Initial conditions
x = np.array([[0], [0]])  # [position, velocity]
P = np.zeros((2, 2))  # Initial covariance

# State transition matrix
F = np.array([[1, dt], [0, 1]])

# Process noise covariance (constant acceleration model)
# Q = σ² * [[Δt⁴/4, Δt³/2], [Δt³/2, Δt²]]
sigma2 = SDg**2
Q = sigma2 * np.array([[dt**4/4, dt**3/2], [dt**3/2, dt**2]])

# Measurements
z = np.array([-2.762e-4, -1.614e-2, -4.453e-2])

# Store results
results = []

for k in range(3):
    print(f"\n=== Iteration k={k} ===")
    
    # Prediction step
    x_minus = F @ x
    P_minus = F @ P @ F.T + Q
    
    print(f"x⁻ = {x_minus.T}")
    print(f"P⁻ =\n{P_minus}")
    
    # Kalman gain
    S = H @ P_minus @ H.T + R  # Innovation covariance
    K = P_minus @ H.T / S
    
    print(f"K = {K.T}")
    
    # Innovation (measurement residual)
    y = z[k] - H @ x_minus
    print(f"y (innovation) = {y}")
    
    # Update step
    x_plus = x_minus + K * y
    P_plus = (np.eye(2) - K @ H) @ P_minus
    
    print(f"x⁺ = {x_plus.T}")
    print(f"P⁺ =\n{P_plus}")
    
    # Store for output
    results.append({
        'k': k,
        'x_minus_pos': x_minus[0, 0],
        'x_minus_vel': x_minus[1, 0],
        'P_minus_00': P_minus[0, 0],
        'P_minus_01': P_minus[0, 1],
        'P_minus_10': P_minus[1, 0],
        'P_minus_11': P_minus[1, 1],
        'K_pos': K[0, 0],
        'K_vel': K[1, 0],
        'x_plus_pos': x_plus[0, 0],
        'x_plus_vel': x_plus[1, 0],
        'P_plus_00': P_plus[0, 0],
        'P_plus_01': P_plus[0, 1],
        'P_plus_10': P_plus[1, 0],
        'P_plus_11': P_plus[1, 1],
    })
    
    # Update state for next iteration
    x = x_plus
    P = P_plus

# Create Excel file
wb = Workbook()
ws = wb.active
ws.title = "Kalman Filter"

# Add title
ws['A1'] = "Kalman Filter Results (k=0,1,2)"
ws['A1'].font = Font(bold=True, size=12)

# Add parameters section
row = 3
ws[f'A{row}'] = "Parameters"
ws[f'A{row}'].font = Font(bold=True)
row += 1
ws[f'A{row}'] = f"dt = {dt} s"
row += 1
ws[f'A{row}'] = f"R = {R} m²"
row += 1
ws[f'A{row}'] = f"SDg = {SDg:.6f} m/s²"
row += 1
ws[f'A{row}'] = f"H = [1 0]"
row += 1

# Add Q matrix
row += 1
ws[f'A{row}'] = "Q ="
ws[f'B{row}'] = Q[0, 0]
ws[f'C{row}'] = Q[0, 1]
row += 1
ws[f'A{row}'] = ""
ws[f'B{row}'] = Q[1, 0]
ws[f'C{row}'] = Q[1, 1]
row += 2

# Add results table
ws[f'A{row}'] = "Results"
ws[f'A{row}'].font = Font(bold=True, size=11)
row += 1

# Headers
headers = ['k', 'x⁻ (pos)', 'x⁻ (vel)', 'P⁻[0,0]', 'P⁻[0,1]', 'P⁻[1,0]', 'P⁻[1,1]', 
           'K (pos)', 'K (vel)', 'x⁺ (pos)', 'x⁺ (vel)', 'P⁺[0,0]', 'P⁺[0,1]', 'P⁺[1,0]', 'P⁺[1,1]']

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

row += 1

# Data rows
for result in results:
    ws.cell(row=row, column=1).value = result['k']
    ws.cell(row=row, column=2).value = result['x_minus_pos']
    ws.cell(row=row, column=3).value = result['x_minus_vel']
    ws.cell(row=row, column=4).value = result['P_minus_00']
    ws.cell(row=row, column=5).value = result['P_minus_01']
    ws.cell(row=row, column=6).value = result['P_minus_10']
    ws.cell(row=row, column=7).value = result['P_minus_11']
    ws.cell(row=row, column=8).value = result['K_pos']
    ws.cell(row=row, column=9).value = result['K_vel']
    ws.cell(row=row, column=10).value = result['x_plus_pos']
    ws.cell(row=row, column=11).value = result['x_plus_vel']
    ws.cell(row=row, column=12).value = result['P_plus_00']
    ws.cell(row=row, column=13).value = result['P_plus_01']
    ws.cell(row=row, column=14).value = result['P_plus_10']
    ws.cell(row=row, column=15).value = result['P_plus_11']
    row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 12
for col in range(2, 16):
    ws.column_dimensions[chr(64 + col)].width = 14

# Save file
output_file = "Kalman_Filter_Results.xlsx"
wb.save(output_file)
print(f"\n✓ Results saved to {output_file}")
